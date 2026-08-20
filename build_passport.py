"""
Build the AMP-GEN passport outputs from the manually validated extraction JSON.
The assessment scan is noisy, so OCR is used as a first pass and the final
records are visually checked against the scanned pages before export.
"""
import argparse, json, os
import openpyxl

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--template", required=True)
    ap.add_argument("--records", required=True)
    ap.add_argument("--outdir", default="output")
    args = ap.parse_args()

    os.makedirs(args.outdir, exist_ok=True)
    with open(args.records, encoding="utf-8") as f:
        records = json.load(f)

    # JSON export
    with open(os.path.join(args.outdir, "passport.json"), "w", encoding="utf-8") as f:
        json.dump(records, f, indent=2, ensure_ascii=False)

    wb = openpyxl.load_workbook(args.template)
    ws = wb["Material Passport"]
    headers = [ws.cell(3, c).value for c in range(1, ws.max_column + 1)]

    for r in range(7, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).value = None

    for r, record in enumerate(records, start=7):
        for c, header in enumerate(headers, start=1):
            ws.cell(r, c).value = record.get(header)

    wb.save(os.path.join(args.outdir, "passport_filled.xlsx"))
    print(f"Wrote {len(records)} records to {args.outdir}")

if __name__ == "__main__":
    main()
